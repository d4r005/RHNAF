#!/usr/bin/env python3
"""
import_excel_data.py — Importa datos históricos del Excel a la API de RHNAF

Importa:
  - CALCIO, PVC, CARTON → /api/v1/recepcion-mp
  - TARIMAS → /api/v1/almacen/tarimas
  - CONTENEDORES CHINA → /api/v1/embarques/contenedores-china
  - SELLOS → /api/v1/almacen/sellos
  - BITACORA DE GAS → /api/v1/ehs/gas

Uso:
  python3 import_excel_data.py                    # importa todo
  python3 import_excel_data.py --dry-run           # solo cuenta, no sube nada
  python3 import_excel_data.py --hoja "CALCIO "    # solo una hoja
  python3 import_excel_data.py --url http://localhost:7860  # URL custom
"""

import openpyxl
import requests
import sys
import argparse
from datetime import datetime

BACKEND_URL = "https://d4r005-rhnaf-industrial.hf.space"

def safe_str(val):
    if val is None: return ""
    return str(val).strip()

def safe_date(val):
    if val is None: return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return safe_str(val)

def import_recepcion_mp(ws, tipo, url, dry_run):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = safe_date(row[0]) if len(row) > 0 else ""
        desc = safe_str(row[1]) if len(row) > 1 else ""
        prov = safe_str(row[2]) if len(row) > 2 else ""
        cant = safe_str(row[3]) if len(row) > 3 else ""
        folio = safe_str(row[4]) if len(row) > 4 else ""
        
        if not fecha and not desc and not cant:
            continue
        
        # PVC tiene estructura diferente: FECHA, PROVEDOR, COSTALES, KG, TOTAL, FOLIO
        if tipo == "PVC":
            prov = safe_str(row[1]) if len(row) > 1 else ""
            cant = f"{safe_str(row[2])} costales / {safe_str(row[3])} kg" if len(row) > 3 else ""
            folio = safe_str(row[5]) if len(row) > 5 else ""
            desc = "PVC"
        
        payload = {
            "fecha": fecha,
            "tipo": tipo,
            "descripcion": desc,
            "proveedor": prov,
            "cantidad": cant,
            "unidad": "kg" if tipo in ("CALCIO", "PVC") else "pza",
            "folio": folio,
            "notas": ""
        }
        if not dry_run:
            try:
                r = requests.post(f"{url}/api/v1/recepcion-mp", json=payload, timeout=15)
                if r.status_code == 200: count += 1
            except: pass
        else:
            count += 1
    return count

def import_tarimas(ws, url, dry_run):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        compania = safe_str(row[0]) if len(row) > 0 else ""
        fecha_llegada = safe_date(row[1]) if len(row) > 1 else ""
        folio = safe_str(row[2]) if len(row) > 2 else ""
        cantidad = safe_str(row[3]) if len(row) > 3 else ""
        medidas = safe_str(row[4]) if len(row) > 4 else ""
        rechazadas = safe_str(row[5]) if len(row) > 5 else ""
        aceptable = safe_str(row[6]) if len(row) > 6 else ""
        fecha_regreso = safe_date(row[7]) if len(row) > 7 else ""
        
        if not folio and not compania:
            continue
        
        payload = {
            "compania": compania,
            "fechaLlegada": fecha_llegada,
            "folio": folio,
            "cantidad": cantidad,
            "medidas": medidas,
            "rechazadas": rechazadas,
            "cantidadAceptable": aceptable,
            "fechaRegreso": fecha_regreso
        }
        if not dry_run:
            try:
                r = requests.post(f"{url}/api/v1/almacen/tarimas", json=payload, timeout=15)
                if r.status_code == 200: count += 1
            except: pass
        else:
            count += 1
    return count

def import_contenedores_china(ws, url, dry_run):
    count = 0
    last_fecha = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = safe_date(row[0]) if len(row) > 0 else ""
        if fecha: last_fecha = fecha
        else: fecha = last_fecha  # usar fecha anterior si está vacía
        
        codigo = safe_str(row[1]) if len(row) > 1 else ""
        nombre = safe_str(row[2]) if len(row) > 2 else ""
        spec = safe_str(row[3]) if len(row) > 3 else ""
        modelo = safe_str(row[4]) if len(row) > 4 else ""
        cantidad = safe_str(row[5]) if len(row) > 5 else ""
        unidad = safe_str(row[6]) if len(row) > 6 else ""
        
        if not codigo:
            continue
        
        payload = {
            "fecha": fecha,
            "codigoProducto": codigo,
            "nombre": nombre,
            "especificacion": spec,
            "modelo": modelo,
            "cantidad": cantidad,
            "unidad": unidad
        }
        if not dry_run:
            try:
                r = requests.post(f"{url}/api/v1/embarques/contenedores-china", json=payload, timeout=15)
                if r.status_code == 200: count += 1
            except: pass
        else:
            count += 1
    return count

def import_sellos(ws, url, dry_run):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = safe_date(row[0]) if len(row) > 0 else ""
        inicial = safe_str(row[1]) if len(row) > 1 else ""
        final = safe_str(row[3]) if len(row) > 3 else ""
        cantidad = safe_str(row[4]) if len(row) > 4 else ""
        
        if not inicial and not final:
            continue
        
        payload = {
            "fecha": fecha,
            "numeroInicial": inicial,
            "numeroFinal": final,
            "cantidad": cantidad
        }
        if not dry_run:
            try:
                r = requests.post(f"{url}/api/v1/almacen/sellos", json=payload, timeout=15)
                if r.status_code == 200: count += 1
            except: pass
        else:
            count += 1
    return count

def import_gas(ws, url, dry_run):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = safe_date(row[0]) if len(row) > 0 else ""
        cantidad_texto = safe_str(row[1]) if len(row) > 1 else ""
        dia = safe_str(row[2]) if len(row) > 2 else ""
        
        if not fecha and not cantidad_texto:
            continue
        
        payload = {
            "fecha": fecha,
            "cantidadTexto": cantidad_texto,
            "diaCarga": dia
        }
        if not dry_run:
            try:
                r = requests.post(f"{url}/api/v1/ehs/gas", json=payload, timeout=15)
                if r.status_code == 200: count += 1
            except: pass
        else:
            count += 1
    return count

def main():
    parser = argparse.ArgumentParser(description="Importa datos del Excel a RHNAF")
    parser.add_argument("--dry-run", action="store_true", help="Solo contar, no subir")
    parser.add_argument("--hoja", type=str, default=None, help="Solo importar esta hoja")
    parser.add_argument("--url", type=str, default=BACKEND_URL, help="URL del backend")
    parser.add_argument("--file", type=str, default="incoming_files/0c12adf36_inventarioporkgpzaym2.xlsx")
    args = parser.parse_args()
    
    print(f"Excel: {args.file}")
    print(f"Backend: {args.url}")
    print(f"Dry run: {args.dry_run}")
    print()
    
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    
    imports = {
        "CALCIO ":      lambda ws: import_recepcion_mp(ws, "CALCIO", args.url, args.dry_run),
        "PVC":          lambda ws: import_recepcion_mp(ws, "PVC", args.url, args.dry_run),
        "CARTON":       lambda ws: import_recepcion_mp(ws, "CARTON", args.url, args.dry_run),
        "TARIMAS ":     lambda ws: import_tarimas(ws, args.url, args.dry_run),
        " CONTANADOR QUE LLEGA DE CHINA ": lambda ws: import_contenedores_china(ws, args.url, args.dry_run),
        "sellos en stock": lambda ws: import_sellos(ws, args.url, args.dry_run),
        "BITACORA DE GAS": lambda ws: import_gas(ws, args.url, args.dry_run),
    }
    
    total = 0
    for sheet_name, func in imports.items():
        if args.hoja and args.hoja != sheet_name:
            continue
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Hoja '{sheet_name}' no encontrada")
            continue
        ws = wb[sheet_name]
        print(f"Importando '{sheet_name}' ({ws.max_row - 1} filas)...", end=" ")
        count = func(ws)
        total += count
        print(f"✓ {count} registros {'(dry-run)' if args.dry_run else 'subidos'}")
    
    print(f"\nTotal: {total} registros {'(dry-run)' if args.dry_run else 'importados'}")
    wb.close()

if __name__ == "__main__":
    main()
